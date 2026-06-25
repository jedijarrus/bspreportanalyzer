"""Tests für app.parser — Einlesen + Normalisieren der RVKU-KI-xlsx."""
import datetime as dt

import openpyxl
import pytest

from app import parser, schema


def test_report_date_aus_dateiname(make_report):
    path = make_report(rows=3, report_date=dt.datetime(2026, 6, 25, 8, 43, 50))
    result = parser.parse_report(path)
    assert result.report_date == dt.datetime(2026, 6, 25, 8, 43, 50)


def test_dateiname_wird_gespeichert(make_report):
    path = make_report(rows=3)
    result = parser.parse_report(path)
    assert result.filename == path.name


def test_zeilenanzahl(make_report):
    path = make_report(rows=17)
    result = parser.parse_report(path)
    assert len(result.rows) == 17


def test_jede_zeile_hat_alle_felder(make_report):
    path = make_report(rows=5)
    result = parser.parse_report(path)
    for row in result.rows:
        assert set(row.keys()) == set(schema.FIELDS)


def test_datetime_felder_bleiben_datetime(make_report):
    path = make_report(rows=5)
    result = parser.parse_report(path)
    assert isinstance(result.rows[0]["vertragsbeginn"], dt.datetime)
    assert isinstance(result.rows[0]["bindefristende"], dt.datetime)


def test_backslash_platzhalter_wird_none(make_report, tmp_path):
    # Telekom-Export nutzt '\' als Leer-Platzhalter -> muss None werden
    path = make_report(rows=2)
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    col = schema.RAW_COLUMNS.index("MultiSIM-Karten-/Profilnummer 4") + 1
    ws.cell(row=2, column=col, value="\\")
    wb.save(path)

    result = parser.parse_report(path)
    assert result.rows[0]["multisim_karten_profilnummer_4"] is None


def test_leerstring_wird_none(make_report, tmp_path):
    path = make_report(rows=2)
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    col = schema.RAW_COLUMNS.index("Sperren") + 1
    ws.cell(row=2, column=col, value="")
    wb.save(path)

    result = parser.parse_report(path)
    assert result.rows[0]["sperren"] is None


def test_leere_datenzeilen_werden_uebersprungen(make_report, tmp_path):
    path = make_report(rows=3)
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    ws.append([None] * len(schema.RAW_COLUMNS))  # komplett leere Zeile
    wb.save(path)

    result = parser.parse_report(path)
    assert len(result.rows) == 3


def test_unbekannter_dateiname_report_date_none(make_report):
    path = make_report(rows=2, name="komischer_name.xlsx")
    result = parser.parse_report(path)
    assert result.report_date is None
